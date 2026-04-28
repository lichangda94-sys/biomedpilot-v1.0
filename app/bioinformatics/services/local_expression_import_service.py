from __future__ import annotations

import csv
import json
import re
import zipfile
from dataclasses import asdict
from datetime import datetime, timezone
from pathlib import Path
from xml.etree import ElementTree
from uuid import uuid4

from app.bioinformatics.models.expression_import import ExpressionImportResult
from app.shared.data_center.service import DataCenter
from app.shared.storage import default_storage_root
from app.shared.task_center.service import TaskCenter, TaskRecord, TaskStatus, TaskType


GENE_COLUMN_KEYWORDS = ("gene_symbol", "genesymbol", "probe_id", "gene", "symbol", "probe", "id", "ensembl")
SAMPLE_COLUMN_KEYWORDS = ("gsm", "tcga", "srr", "sample", "control", "tumor", "normal", "treat", "drug")
SUPPORTED_EXTENSIONS = {".csv", ".tsv", ".txt", ".xlsx"}
MISSING_VALUES = {"", "na", "n/a", "nan", "null", "none"}


class LocalExpressionImportService:
    def __init__(
        self,
        *,
        task_center: TaskCenter | None = None,
        data_center: DataCenter | None = None,
        storage_root: Path | None = None,
    ) -> None:
        self._task_center = task_center or TaskCenter.default()
        self._data_center = data_center or DataCenter.default()
        self._storage_root = storage_root or default_storage_root()

    def import_expression_matrix(self, *, project_id: str, source_path: str) -> ExpressionImportResult:
        task = self._start_task(project_id=project_id, source_path=source_path)
        validation_error = self._validate(source_path)
        if validation_error is not None:
            result = ExpressionImportResult(
                success=False,
                source_path=source_path,
                source_type="",
                row_count=0,
                column_count=0,
                candidate_gene_columns=[],
                candidate_sample_columns=[],
                numeric_sample_column_count=0,
                missing_value_rate=0.0,
                output_path="",
                warnings=[],
                message=validation_error,
            )
            self._finish_task(task, result)
            return result

        path = Path(source_path).expanduser().resolve()
        source_type = path.suffix.lower().lstrip(".")
        try:
            headers, rows = self._read_table(path)
            summary = self._summarize(headers, rows)
            asset_id = f"expression-matrix-{uuid4().hex[:12]}"
            summary_path, manifest_path = self._write_outputs(
                project_id=project_id,
                source_path=path,
                source_type=source_type,
                result=summary,
                asset_id=asset_id,
            )
            result = ExpressionImportResult(
                success=True,
                source_path=str(path),
                source_type=source_type,
                row_count=summary.row_count,
                column_count=summary.column_count,
                candidate_gene_columns=summary.candidate_gene_columns,
                candidate_sample_columns=summary.candidate_sample_columns,
                numeric_sample_column_count=summary.numeric_sample_column_count,
                missing_value_rate=summary.missing_value_rate,
                output_path=str(manifest_path),
                warnings=summary.warnings,
                message=(
                    f"本地表达矩阵导入预检完成：{summary.row_count} 行，"
                    f"{summary.column_count} 列，识别 {summary.numeric_sample_column_count} 个数值样本列。"
                ),
                details={
                    "raw_file_modified": False,
                    "normalization_executed": False,
                    "summary_path": str(summary_path),
                    "manifest_path": str(manifest_path),
                },
                asset_id=asset_id,
                summary_path=str(summary_path),
                manifest_path=str(manifest_path),
                gene_id_column_candidates=summary.gene_id_column_candidates,
                selected_gene_id_column=None,
                sample_expression_column_candidates=summary.sample_expression_column_candidates,
                numeric_column_count=summary.numeric_column_count,
                numeric_column_ratio=summary.numeric_column_ratio,
                missing_value_summary=summary.missing_value_summary,
                duplicate_gene_id_count=summary.duplicate_gene_id_count,
                non_numeric_columns=summary.non_numeric_columns,
                is_expression_matrix_suitable=summary.is_expression_matrix_suitable,
                errors=summary.errors,
            )
            self._data_center.register_asset(
                project_id=project_id,
                module="bioinformatics",
                data_type="expression_matrix",
                source_path=str(path),
                output_path=str(manifest_path),
                status="available",
                data_id=asset_id,
            )
            self._finish_task(task, result)
            return result
        except ImportError as exc:
            result = self._failure(path, source_type, "读取 XLSX 需要安装 openpyxl，请安装后重试。", {"error": str(exc)})
            self._finish_task(task, result)
            return result
        except Exception as exc:
            result = self._failure(path, source_type, "表达矩阵导入失败，请确认文件是带表头的表格文件。", {"error": str(exc)})
            self._finish_task(task, result)
            return result

    def _validate(self, source_path: str) -> str | None:
        if not source_path.strip():
            return "请输入本地表达矩阵文件路径。"
        path = Path(source_path).expanduser()
        if not path.exists():
            return "表达矩阵文件不存在，请检查路径。"
        if not path.is_file():
            return "表达矩阵路径需要指向一个文件。"
        if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            return "暂不支持该文件格式，请使用 CSV、TSV、TXT 或 XLSX 文件。"
        return None

    def _read_table(self, path: Path) -> tuple[list[str], list[list[object]]]:
        if path.suffix.lower() == ".xlsx":
            return self._read_xlsx(path)
        delimiter = "," if path.suffix.lower() == ".csv" else "\t"
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle, delimiter=delimiter)
            rows = [row for row in reader]
        if not rows:
            raise ValueError("empty_table")
        headers = [str(value).strip() for value in rows[0]]
        return headers, rows[1:]

    def _read_xlsx(self, path: Path) -> tuple[list[str], list[list[object]]]:
        try:
            from openpyxl import load_workbook
        except Exception:
            return self._read_xlsx_fallback(path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        workbook.close()
        if not rows:
            raise ValueError("empty_table")
        headers = ["" if value is None else str(value).strip() for value in rows[0]]
        data_rows = [list(row) for row in rows[1:]]
        return headers, data_rows

    def _read_xlsx_fallback(self, path: Path) -> tuple[list[str], list[list[object]]]:
        with zipfile.ZipFile(path) as archive:
            names = set(archive.namelist())
            if "xl/worksheets/sheet1.xml" not in names:
                raise ImportError("openpyxl is required for XLSX files with unsupported workbook layout")
            shared_strings = self._read_xlsx_shared_strings(archive, names)
            worksheet = ElementTree.fromstring(archive.read("xl/worksheets/sheet1.xml"))
            rows: list[list[object]] = []
            for row in worksheet.findall(".//{*}sheetData/{*}row"):
                values_by_column: dict[int, object] = {}
                for cell in row.findall("{*}c"):
                    reference = str(cell.attrib.get("r", ""))
                    column_index = _xlsx_column_index(reference)
                    values_by_column[column_index] = self._xlsx_cell_value(cell, shared_strings)
                if values_by_column:
                    max_index = max(values_by_column)
                    rows.append([values_by_column.get(index, "") for index in range(max_index + 1)])
            if not rows:
                raise ValueError("empty_table")
            headers = [str(value).strip() for value in rows[0]]
            return headers, rows[1:]

    def _read_xlsx_shared_strings(self, archive: zipfile.ZipFile, names: set[str]) -> list[str]:
        if "xl/sharedStrings.xml" not in names:
            return []
        root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
        strings: list[str] = []
        for item in root.findall("{*}si"):
            text_parts = [node.text or "" for node in item.findall(".//{*}t")]
            strings.append("".join(text_parts))
        return strings

    def _xlsx_cell_value(self, cell: ElementTree.Element, shared_strings: list[str]) -> object:
        cell_type = cell.attrib.get("t", "")
        if cell_type == "inlineStr":
            return "".join(node.text or "" for node in cell.findall(".//{*}t"))
        value_node = cell.find("{*}v")
        if value_node is None or value_node.text is None:
            return ""
        raw_value = value_node.text
        if cell_type == "s":
            try:
                return shared_strings[int(raw_value)]
            except (ValueError, IndexError):
                return ""
        try:
            numeric = float(raw_value)
        except ValueError:
            return raw_value
        return int(numeric) if numeric.is_integer() else numeric

    def _summarize(self, headers: list[str], rows: list[list[object]]) -> ExpressionImportResult:
        normalized_headers = [header.strip() or f"column_{index + 1}" for index, header in enumerate(headers)]
        row_count = len(rows)
        column_count = len(normalized_headers)
        gene_columns = [
            header
            for header in normalized_headers
            if any(keyword in _normalize_name(header) for keyword in GENE_COLUMN_KEYWORDS)
        ]
        profiles = {
            header: self._column_profile(rows, index)
            for index, header in enumerate(normalized_headers)
        }
        numeric_columns = [
            header
            for header in normalized_headers
            if header not in gene_columns
            and profiles[header]["numeric_count"] > 0
            and profiles[header]["numeric_ratio"] >= 0.8
        ]
        pattern_columns = [
            header
            for header in normalized_headers
            if header not in gene_columns
            and any(keyword in _normalize_name(header) for keyword in SAMPLE_COLUMN_KEYWORDS)
        ]
        sample_columns = _dedupe([*numeric_columns, *pattern_columns])
        sample_expression_columns = list(numeric_columns)
        missing_value_summary = self._missing_value_summary(rows, normalized_headers, sample_expression_columns)
        missing_value_rate = float(missing_value_summary.get("missing_value_rate", 0.0))
        duplicate_gene_id_count = self._duplicate_gene_id_count(rows, normalized_headers, gene_columns)
        non_numeric_columns = self._non_numeric_columns(normalized_headers, gene_columns, sample_expression_columns, profiles)
        non_gene_column_count = max(column_count - len(gene_columns), 0)
        numeric_column_ratio = round(len(sample_expression_columns) / non_gene_column_count, 4) if non_gene_column_count else 0.0
        errors = self._errors(
            row_count=row_count,
            column_count=column_count,
            gene_columns=gene_columns,
            numeric_sample_column_count=len(sample_expression_columns),
        )
        warnings = self._warnings(
            row_count=row_count,
            column_count=column_count,
            gene_columns=gene_columns,
            numeric_sample_column_count=len(sample_expression_columns),
            missing_value_rate=missing_value_rate,
            duplicate_gene_id_count=duplicate_gene_id_count,
            non_numeric_columns=non_numeric_columns,
        )
        return ExpressionImportResult(
            success=True,
            source_path="",
            source_type="",
            row_count=row_count,
            column_count=column_count,
            candidate_gene_columns=gene_columns,
            candidate_sample_columns=sample_columns,
            numeric_sample_column_count=len(sample_expression_columns),
            missing_value_rate=missing_value_rate,
            output_path="",
            warnings=warnings,
            message="summary",
            gene_id_column_candidates=gene_columns,
            selected_gene_id_column=None,
            sample_expression_column_candidates=sample_expression_columns,
            numeric_column_count=len(sample_expression_columns),
            numeric_column_ratio=numeric_column_ratio,
            missing_value_summary=missing_value_summary,
            duplicate_gene_id_count=duplicate_gene_id_count,
            non_numeric_columns=non_numeric_columns,
            is_expression_matrix_suitable=not errors,
            errors=errors,
        )

    def _column_profile(self, rows: list[list[object]], column_index: int) -> dict[str, object]:
        missing_count = 0
        non_missing_count = 0
        numeric_count = 0
        non_numeric_examples: list[str] = []
        for row in rows:
            value = row[column_index] if column_index < len(row) else ""
            if _is_missing(value):
                missing_count += 1
                continue
            non_missing_count += 1
            try:
                float(str(value).strip())
            except ValueError:
                if len(non_numeric_examples) < 3:
                    non_numeric_examples.append(str(value))
                continue
            numeric_count += 1
        numeric_ratio = round(numeric_count / non_missing_count, 4) if non_missing_count else 0.0
        return {
            "missing_count": missing_count,
            "non_missing_count": non_missing_count,
            "numeric_count": numeric_count,
            "non_numeric_count": max(non_missing_count - numeric_count, 0),
            "numeric_ratio": numeric_ratio,
            "non_numeric_examples": non_numeric_examples,
        }

    def _missing_value_summary(self, rows: list[list[object]], headers: list[str], sample_columns: list[str]) -> dict[str, object]:
        if not rows or not sample_columns:
            return {
                "total_cells": 0,
                "missing_cells": 0,
                "missing_value_rate": 0.0,
                "by_column": {},
            }
        indexes = [headers.index(column) for column in sample_columns]
        total = len(rows) * len(indexes)
        missing = 0
        by_column: dict[str, dict[str, object]] = {}
        for column, index in zip(sample_columns, indexes):
            column_missing = 0
            for row in rows:
                value = row[index] if index < len(row) else ""
                if _is_missing(value):
                    column_missing += 1
            by_column[column] = {
                "missing_count": column_missing,
                "missing_value_rate": round(column_missing / len(rows), 4) if rows else 0.0,
            }
        for row in rows:
            for index in indexes:
                value = row[index] if index < len(row) else ""
                if _is_missing(value):
                    missing += 1
        return {
            "total_cells": total,
            "missing_cells": missing,
            "missing_value_rate": round(missing / total, 4) if total else 0.0,
            "by_column": by_column,
        }

    def _duplicate_gene_id_count(self, rows: list[list[object]], headers: list[str], gene_columns: list[str]) -> int:
        if not gene_columns:
            return 0
        index = headers.index(gene_columns[0])
        counts: dict[str, int] = {}
        for row in rows:
            value = row[index] if index < len(row) else ""
            if _is_missing(value):
                continue
            key = str(value).strip()
            counts[key] = counts.get(key, 0) + 1
        return sum(count - 1 for count in counts.values() if count > 1)

    def _non_numeric_columns(
        self,
        headers: list[str],
        gene_columns: list[str],
        sample_expression_columns: list[str],
        profiles: dict[str, dict[str, object]],
    ) -> list[str]:
        non_numeric: list[str] = []
        for header in headers:
            if header in gene_columns or header in sample_expression_columns:
                continue
            profile = profiles[header]
            if profile["non_missing_count"] and profile["numeric_ratio"] < 0.8:
                non_numeric.append(header)
        return non_numeric

    def _errors(
        self,
        *,
        row_count: int,
        column_count: int,
        gene_columns: list[str],
        numeric_sample_column_count: int,
    ) -> list[str]:
        errors: list[str] = []
        if row_count < 1 or column_count < 2:
            errors.append("文件行列数异常，无法作为表达矩阵资产。")
        if not gene_columns:
            errors.append("未识别到 gene/probe/id 列。")
        if numeric_sample_column_count < 2:
            errors.append("数值样本表达列少于 2 个。")
        return errors

    def _warnings(
        self,
        *,
        row_count: int,
        column_count: int,
        gene_columns: list[str],
        numeric_sample_column_count: int,
        missing_value_rate: float,
        duplicate_gene_id_count: int,
        non_numeric_columns: list[str],
    ) -> list[str]:
        warnings: list[str] = []
        if not gene_columns:
            warnings.append("没有识别到 gene/probe 列，请在数据资产确认步骤手动选择。")
        if numeric_sample_column_count < 2:
            warnings.append("数值样本列太少，后续分组和差异分析可能无法运行。")
        if missing_value_rate >= 0.2:
            warnings.append("缺失值比例较高，后续需要清洗或过滤。")
        if duplicate_gene_id_count:
            warnings.append(f"检测到 {duplicate_gene_id_count} 条重复 gene/probe/id，后续清洗需要聚合处理。")
        if non_numeric_columns:
            warnings.append(f"以下非数值列未作为表达样本列：{', '.join(non_numeric_columns)}。")
        if row_count < 1 or column_count < 2:
            warnings.append("文件行列数异常，请确认表格包含基因列和样本列。")
        return warnings

    def _write_outputs(
        self,
        *,
        project_id: str,
        source_path: Path,
        source_type: str,
        result: ExpressionImportResult,
        asset_id: str,
    ) -> tuple[Path, Path]:
        output_dir = self._storage_root / "projects" / project_id / "bioinformatics" / "expression_import" / f"import-{uuid4().hex[:12]}"
        output_dir.mkdir(parents=True, exist_ok=True)
        created_at = datetime.now(timezone.utc).isoformat()
        summary_path = output_dir / "expression_matrix_import_summary.json"
        manifest_path = output_dir / "expression_matrix_asset_manifest.json"
        summary_payload = {
            "project_id": project_id,
            "module": "bioinformatics",
            "data_type": "expression_matrix",
            "asset_id": asset_id,
            "source_path": str(source_path),
            "source_type": source_type,
            "created_at": created_at,
            "status": "ready_for_asset_confirmation",
            "raw_file_modified": False,
            "normalization_executed": False,
            "summary": asdict(result) | {"source_path": str(source_path), "source_type": source_type},
        }
        manifest_payload = {
            "project_id": project_id,
            "module": "bioinformatics",
            "data_type": "expression_matrix",
            "asset_id": asset_id,
            "asset_type": "expression_matrix",
            "source_file": str(source_path),
            "source_path": str(source_path),
            "file_format": source_type,
            "source_type": source_type,
            "row_count": result.row_count,
            "column_count": result.column_count,
            "gene_id_column_candidates": result.gene_id_column_candidates,
            "selected_gene_id_column": None,
            "sample_column_candidates": result.sample_expression_column_candidates,
            "numeric_column_count": result.numeric_column_count,
            "numeric_column_ratio": result.numeric_column_ratio,
            "missing_value_summary": result.missing_value_summary,
            "duplicate_gene_id_count": result.duplicate_gene_id_count,
            "non_numeric_columns": result.non_numeric_columns,
            "created_at": created_at,
            "status": "ready_for_asset_confirmation" if result.is_expression_matrix_suitable else "needs_review",
            "is_expression_matrix_suitable": result.is_expression_matrix_suitable,
            "warnings": result.warnings,
            "errors": result.errors,
            "raw_file_modified": False,
            "normalization_executed": False,
            "summary_path": str(summary_path),
            "summary": summary_payload["summary"],
        }
        summary_path.write_text(json.dumps(summary_payload, ensure_ascii=False, indent=2), encoding="utf-8")
        manifest_path.write_text(json.dumps(manifest_payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return summary_path, manifest_path

    def _failure(self, path: Path, source_type: str, message: str, details: dict[str, object]) -> ExpressionImportResult:
        return ExpressionImportResult(
            success=False,
            source_path=str(path),
            source_type=source_type,
            row_count=0,
            column_count=0,
            candidate_gene_columns=[],
            candidate_sample_columns=[],
            numeric_sample_column_count=0,
            missing_value_rate=0.0,
            output_path="",
            warnings=[],
            message=message,
            details=details,
            errors=[message],
        )

    def _start_task(self, *, project_id: str, source_path: str) -> TaskRecord:
        now = datetime.now(timezone.utc).isoformat()
        return self._task_center.register_task(
            task_id=f"task-{uuid4().hex[:12]}",
            task_type=TaskType.IMPORT,
            module="bioinformatics",
            title="Local Expression Matrix Import",
            project_id=project_id,
            status=TaskStatus.RUNNING,
            started_at=now,
            summary=f"local_expression_import: {source_path}" if source_path else "local_expression_import: waiting for file path",
        )

    def _finish_task(self, task: TaskRecord, result: ExpressionImportResult) -> None:
        now = datetime.now(timezone.utc).isoformat()
        self._task_center.save_task(
            TaskRecord(
                task_id=task.task_id,
                task_type=task.task_type,
                status=TaskStatus.COMPLETED if result.success else TaskStatus.FAILED,
                module=task.module,
                title=task.title,
                created_at=task.created_at,
                updated_at=now,
                project_id=task.project_id,
                started_at=task.started_at,
                finished_at=now,
                summary=f"local_expression_import: {result.message}",
                error_message="" if result.success else result.message,
            )
        )


def _normalize_name(value: str) -> str:
    return value.strip().lower().replace("-", "_").replace(" ", "_")


def _is_missing(value: object) -> bool:
    if value is None:
        return True
    return str(value).strip().lower() in MISSING_VALUES


def _dedupe(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value not in seen:
            seen.add(value)
            result.append(value)
    return result


def _xlsx_column_index(reference: str) -> int:
    match = re.match(r"([A-Za-z]+)", reference)
    if not match:
        return 0
    index = 0
    for character in match.group(1).upper():
        index = index * 26 + (ord(character) - ord("A") + 1)
    return max(index - 1, 0)
